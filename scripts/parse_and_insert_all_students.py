import os
import re
import json
import zipfile
import openpyxl
import docx

def guess_gender(name):
    n = name.strip().lower()
    last_word = n.split()[-1] if n else ""
    if last_word in ['putri', 'azzahra', 'ayu', 'nisa', 'zahra', 'dewi', 'sari', 'nur', 'sakhi', 'hanun', 'sakina', 'khalisa', 'pustika', 'azalia', 'utami', 'maheswari', 'rahmani', 'wibowo', 'safwana', 'adhisty', 'wimala', 'nistrianto', 'kusuma', 'safira', 'khairunnisa']:
        return 'Perempuan'
    if last_word in ['putra', 'akbar', 'pratama', 'hidayat', 'ramadhan', 'setiawan', 'prasojo', 'nugroho', 'kamal', 'falah', 'wijaya', 'yusuf', 'wibawa', 'santoso', 'dewantoro', 'wiratama', 'barra', 'siddiq', 'natta', 'annam', 'prawiro', 'pradipta', 'wardana', 'solhidar', 'askari', 'abqari']:
        return 'Laki-laki'
    first_word = n.split()[0] if n else ""
    if first_word in ['aisyah', 'anindita', 'alisha', 'arika', 'azzahra', 'bilqis', 'callista', 'dzakira', 'nafia', 'annisa', 'alika', 'kuntum', 'shevahira', 'tyasayu', 'harumi', 'acintya', 'aifriza', 'ainayya', 'alesha', 'kinarayu', 'gendis', 'adzkiya', 'arsyila', 'elshanum', 'qiandra', 'yurike', 'zizi', 'elmira', 'naila', 'luna', 'nawra', 'assyabia', 'kayyisah', 'shaqueena', 'clarisha', 'chayra', 'adifa', 'adreena', 'aisha', 'almahyra', 'alsyakila', 'aqilla', 'aqila', 'arsya', 'ceisya', 'gladysa', 'khayra', 'myeisha', 'queenzy', 'shakeela', 'zianka', 'adiba', 'bellvania', 'khalifa', 'fiona', 'fifi', 'siti', 'dwi', 'tri', 'ratna', 'clemira', 'enzy', 'kaysha', 'khanza', 'mikhayla', 'nadhira', 'nadine', 'najwa', 'revalena']:
        return 'Perempuan'
    return 'Laki-laki'

def clean_name(val_str):
    if not val_str: return ""
    val_str = str(val_str).strip()
    val_str = re.sub(r'^\d+[\.\s\-]+', '', val_str).strip()
    return val_str

def is_valid_name(name):
    if not name or len(name) < 3: return False
    if re.match(r'^[\d\W]+$', name): return False
    bad_kws = ["KOTA", "DAFTAR", "ABSEN", "KELAS", "SEKOLAH", "MADRASAH", "TAHUN", "WALI", "JUMLAH", "MADIUN", "REKAP", "TEMPAT", "TANGGAL", "NAMA", "JENIS", "KETERANGAN", "PEMBAGIAN", "PEMERINTAH", "KEMENTERIAN", "KEPALA", "MADRASAH"]
    if any(kw in name.upper() for kw in bad_kws): return False
    return True

def parse_excel_file(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return parse_plain_text(filepath)
    
    results = []
    seen = set()
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for row in sheet.iter_rows(values_only=True):
            if not row: continue
            row_strs = [str(cell).strip() if cell is not None else "" for cell in row]
            
            gender = None
            for val in row_strs:
                if val.lower() in ['l', 'laki-laki', 'laki', 'l (laki-laki)']: gender = 'Laki-laki'
                elif val.lower() in ['p', 'perempuan', 'p (perempuan)']: gender = 'Perempuan'
            
            name = ""
            for val in row_strs:
                cleaned = clean_name(val)
                if is_valid_name(cleaned):
                    if len(cleaned) > len(name):
                        name = cleaned
            
            if name and name not in seen:
                seen.add(name)
                if not gender: gender = guess_gender(name)
                results.append({"name": name, "gender": gender})
    return results

def parse_plain_text(filepath):
    results = []
    seen = set()
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            cleaned = clean_name(line)
            if is_valid_name(cleaned):
                if cleaned not in seen:
                    seen.add(cleaned)
                    results.append({"name": cleaned, "gender": guess_gender(cleaned)})
    return results

def parse_docx_file(filepath):
    results = []
    seen = set()
    try:
        doc = docx.Document(filepath)
        for table in doc.tables:
            for row in table.rows:
                row_strs = [cell.text.strip() for cell in row.cells]
                gender = None
                for val in row_strs:
                    if val.lower() in ['l', 'laki-laki', 'laki']: gender = 'Laki-laki'
                    elif val.lower() in ['p', 'perempuan']: gender = 'Perempuan'
                
                name = ""
                for cell in row.cells:
                    cleaned = clean_name(cell.text)
                    if is_valid_name(cleaned):
                        if len(cleaned) > len(name):
                            name = cleaned
                if name and name not in seen:
                    seen.add(name)
                    if not gender: gender = guess_gender(name)
                    results.append({"name": name, "gender": gender})
    except Exception:
        try:
            with zipfile.ZipFile(filepath) as z:
                xml_content = z.read('word/document.xml').decode('utf-8')
                text_nodes = re.findall(r'<w:t[^>]*>(.*?)</w:t>', xml_content)
                for t in text_nodes:
                    cleaned = clean_name(t)
                    if is_valid_name(cleaned) and cleaned not in seen:
                        seen.add(cleaned)
                        results.append({"name": cleaned, "gender": guess_gender(cleaned)})
        except Exception:
            pass
    return results

def load_fallback_json():
    if os.path.exists("scripts/final_students.json"):
        with open("scripts/final_students.json") as f:
            data = json.load(f)
            clean_data = {}
            for c_name, st_list in data.items():
                valid = []
                seen = set()
                for s in st_list:
                    c_n = clean_name(s['name'])
                    if is_valid_name(c_n) and c_n not in seen:
                        seen.add(c_n)
                        valid.append({"name": c_n, "gender": s.get('gender', guess_gender(c_n))})
                clean_data[c_name] = valid
            return clean_data
    return {}

def main():
    data_dir = "data"
    parsed_classes = {}
    fallback = load_fallback_json()
    
    file_map = {
        "ABSEN 1A 2026-2027.docx": "Kelas 1A",
        "Absen Kelas 1B.pdf": "Kelas 1B",
        "ABSEN 1C 2627.pdf": "Kelas 1C",
        "ABSEN 2A 2026_2027.xlsx": "Kelas 2A",
        "ABSEN 2B 2627.xlsx": "Kelas 2B",
        "ABSENSI 2C 2026-2027.docx": "Kelas 2C",
        "Data siswa 3C 26-27.docx": "Kelas 3C",
        "Data siswa 4A 2026-2027.docx": "Kelas 4A",
        "ABSENSI KELAS 4B 2627.xlsx": "Kelas 4B",
        "Daftar Siswa Kelas 4C  20262027.docx": "Kelas 4C",
        "Absensi Siswa Kelas 5A 2627.xlsx": "Kelas 5A",
        "ABSENSI KELAS 5B 2526 - Copy.xlsx": "Kelas 5B",
        "DAFTAR PESERTA DIDIK KELAS V-C_T.A 2026-2027.xlsx": "Kelas 5C",
        "ABSEN SISWA 5D 2026 2027.xlsx": "Kelas 5D",
        "Data kelas 6A.xlsx": "Kelas 6A",
        "DAFTAR NAMA KELAS 6B 2026- 2027.xlsx": "Kelas 6B",
        "ABSENSI 6C 2026-2027.xlsx": "Kelas 6C",
        "ABSEN 6D BARU 2026.xlsx": "Kelas 6D"
    }

    for fname, target_class in file_map.items():
        fpath = os.path.join(data_dir, fname)
        students = []
        if os.path.exists(fpath):
            ext = fname.split('.')[-1].lower()
            if ext == 'xlsx':
                students = parse_excel_file(fpath)
            elif ext == 'docx':
                students = parse_docx_file(fpath)
        
        if not students and target_class in fallback:
            students = fallback[target_class]
            
        print(f"[{target_class}] Final count: {len(students)} students")
        parsed_classes[target_class] = students

    with open("scripts/all_classes_parsed.json", "w") as out:
        json.dump(parsed_classes, out, indent=2)

if __name__ == "__main__":
    main()

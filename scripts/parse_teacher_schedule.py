import openpyxl
import os
import re

f1 = "data/JADWAL 2627 UPDATE.xlsx"
f2 = "data/JADWAL KELAS 1-6 (2).xlsx"

def scan_file(filepath):
    print(f"\n==========================================")
    print(f"PARSING FILE: {filepath}")
    print(f"==========================================")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    homerooms = {}
    teacher_codes = {}
    class_subjects = {}

    for sname in wb.sheetnames:
        sheet = wb[sname]
        
        # Check sheet for Wali Kelas title
        for row in sheet.iter_rows(values_only=True):
            if not row: continue
            row_strs = [str(c).strip() if c is not None else "" for c in row]
            row_text = " ".join(row_strs)
            
            # Find Wali Kelas
            if "WALI KELAS" in row_text.upper():
                for idx, cell_text in enumerate(row_strs):
                    if "WALI KELAS" in cell_text.upper():
                        # Wali kelas name might be in this cell or next cell(s)
                        name_candidate = " ".join([c for c in row_strs[idx:] if c and "WALI KELAS" not in c.upper() and c != ":"])
                        if name_candidate:
                            homerooms[sname] = name_candidate
            
            # Find teacher code table if present (e.g. "1 = ...", "KODE GURU", etc.)
            if "KODE GURU" in row_text.upper() or "DAFTAR GURU" in row_text.upper() or "NAMA GURU" in row_text.upper():
                print(f" Found Guru Code Table header in sheet [{sname}]: {row_strs}")

        # Scan for subjects and teacher codes in the schedule table
        subjects_found = set()
        for row in sheet.iter_rows(values_only=True):
            if not row: continue
            row_strs = [str(c).strip() if c is not None else "" for c in row]
            for i in range(len(row_strs) - 1):
                cell_a = row_strs[i]
                cell_b = row_strs[i+1]
                # If cell_a looks like subject name and cell_b looks like code/initial
                if cell_a and len(cell_a) > 2 and cell_a.upper() not in ["SENIN", "SELASA", "RABU", "KAMIS", "JUMAT", "SHALAT DUHA", "ISTIRAHAT", "KOTA MADIUN", "JADWAL PELAJARAN", "TAHUN AJARAN"]:
                    if re.match(r'^[A-Za-z0-9\,\sT]+$', cell_b) and len(cell_b) <= 15:
                        subjects_found.add((cell_a, cell_b))
        
        class_subjects[sname] = list(subjects_found)

    print("\n--- HOMEROOM TEACHERS (WALI KELAS) FOUND ---")
    for c_name, teacher in sorted(homerooms.items()):
        print(f" * {c_name}: {teacher}")
        
    return homerooms, class_subjects

scan_file(f1)
scan_file(f2)

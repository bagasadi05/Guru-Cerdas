import sys
import json
import os
import re

def guess_gender(name):
    # simple heuristic
    n = name.strip().lower()
    last_word = n.split()[-1] if n else ""
    if last_word in ['putri', 'azzahra', 'ayu', 'nisa', 'zahra', 'dewi', 'sari', 'nur']:
        return 'Perempuan'
    if last_word in ['putra', 'akbar', 'pratama', 'hidayat', 'ramadhan']:
        return 'Laki-laki'
        
    # check last letter of first name
    first_word = n.split()[0] if n else ""
    if first_word.endswith('a') or first_word.endswith('i'):
        return 'Perempuan'
    return 'Laki-laki'

def parse_excel_sheets(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = {}
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        students = []
        for row in sheet.iter_rows(values_only=True):
            if not row: continue
            row_strs = [str(cell).strip() if cell is not None else "" for cell in row]
            
            # We will ignore birthplace columns when picking the name, but don't skip the whole row!
            
            gender = None
            for val in row_strs:
                if val.lower() in ['l', 'laki-laki', 'laki', 'l (laki-laki)']: gender = 'Laki-laki'
                elif val.lower() in ['p', 'perempuan', 'p (perempuan)']: gender = 'Perempuan'
            
            name = ""
            for val in row_strs:
                if len(val) > 3 and not re.match(r'^[\d\W]+$', val) and val.lower() not in ['l', 'p', 'laki-laki', 'perempuan', 'l (laki-laki)', 'p (perempuan)']:
                    if len(val) > len(name) and "KOTA" not in val.upper() and "," not in val:
                        name = val
            
            if name:
                if not gender:
                    gender = guess_gender(name)
                students.append({"name": name, "gender": gender})
        if students:
            results[sheetname.strip()] = students
    return results

def parse_docx(filepath):
    import docx
    doc = docx.Document(filepath)
    students = []
    for table in doc.tables:
        for row in table.rows:
            row_data = [cell.text.strip().lower() for cell in row.cells]
            gender = None
            for val in row_data:
                if val in ['l', 'laki-laki', 'laki']: gender = 'Laki-laki'
                elif val in ['p', 'perempuan']: gender = 'Perempuan'
                
            name = ""
            for cell in row.cells:
                val_str = cell.text.strip()
                if len(val_str) > 3 and not re.match(r'^[\d\W]+$', val_str) and val_str.lower() not in ['l', 'p', 'laki-laki', 'perempuan']:
                    if len(val_str) > len(name):
                        name = val_str
            if name:
                if not gender: gender = guess_gender(name)
                students.append({"name": name, "gender": gender})
    return {"Kelas 3C": students}

def parse_pdf_text_manual():
    # 4A, 4B, 4C parsed from fitz output earlier
    text_4a = """FAIREL ALINZKY ADYATAMA
SATRIO MUHAMMAD ATHOILLAH
AHMAD ABIZAR NAZRIL ASRAF
MAGANI ABYAKSA KARTIKO
KENZIE YAFIQ HAMIZAN CAHYONO
NAFIA SALSABILA AS SYIFA
ANNISA AILA LAURYAN
BILQIS ALVIRA CHRISTANTO
DZAKIRA LUTHFIA HARTATI
ABIYASA MANNAF WIRA DEWANTORO
KEANU SATRIO NUGROHO
ANDRUSHA NAGATA EL FAUZI
MANGGALADIPA YOGANATA
RASYID AKMAL PRASETYA
MAHENDRA ZABRAN IRAWAN
ALIKA NAILA PUTRI
KUNTUM KHOIRO UMMAH
SHEVAHIRA FIRDAUSI
TYASAYU SANTIKA ADIADJANA
MUH GHANI BIRAWA ATMAPUTRA
ALFATHAN ABIYA UTOMO
GIBRAN AVISENNA DARIKH
ZAKARIYA DAHLAN APRIYUDI
ATHALLA ASKA BRILLIAN ERWANTO
KEI SYERA JOSHINTA
LAVANYA MAKAILA HAVIE
ZHAHIRAH ALENA WINDONO
QUEENLYA ANGGRAINI FEBRI VALLENZA
ALISHA KINARIAN"""

    text_4b = """ALTHAF MAULANA SHIDQI
REYNAND AMMAR ADHYASTHA
EL FATIH RAZIQ ABBASY
MUHAMMAD RAYYAN ALFATIH
AVEIRO KURNIA BRATADIKARA
NAJULA RAMADHANIA FITRI
AISYA FAIHA PUTRI NADZIFI
HARUMI AZRINA ISWAHYUNI
ACINTYA HAFIDZAH PARAMUSITA
DZULFIKAR JALALUDIN AKBAR
KENZO GASTIASHH ABHYASA
MUHAMMAD RAFIF YUSRO
MUHAMMAD RASYID IBRAHIM
SAKA BUMI YUSUF
AIFRIZA AQILLA SHAFANA
AINAYYA ZAHRA WIJAYANTI
ALESHA NUHA PUTRI PRANATA
KINARAYU EI PANCANAKA
GENDIS PRAMESTI SARBIHANA
MUHAMMAD ABRISYAM AR RAFIF
FABIAN ARSAKHA WINASIS
NEANDRO RAFARDHANI YUDAWAN
DZAKI ILHAM PRABOWO
FAWWAZ BIKRY ABQORY
ADZKIYA NAFI'AH
ARSYILA MEDINA NAYYARA NAFISA
ELSHANUM PUTRI KELDISA
QIANDRA ASHALINA ZAZAKI
YURIKE"""

    text_4c = """ARSENKA ENZO NATTA
KEENAN AL GHAISAN BASKARA
KIYOSHI CATRA BASWARA
ABDULLAH AZZAM AL KHALIFI
AHLAM DYFI ZULFADLY
FIDELYO ARKANANTA ABQARY
ANNISA NAYFA ADZKIA
ZIZI FARADIBA
ALESHA ZAHRA PRIAMBODO
MUHAMMAD AKMAL ADHYASTA
MUHAMMAD RAYYAN
MUHAMMAD ZAYN SETIAWAN
QAISAR RASYAD AL AISY
SAKHA ATHAR KANSA
ADIBA KHANZA AZZAHRA
BELLVANIA ALLMEERA AZKADINA SYAHIRA
KHALIFA SAHZY PRIMAJAYA
ELMIRA YASMINE MAHESWARI DAELA
NAILA DEA ARDININGRUM
RAYYAN YAZID ABIYASA
ALFATIH ALFARIZKI
HAYDAR ARYASATYA AVERUSY
ALFARO ACHIERA LUCKY SETIAWAN
RAFIF MUHAMMAD SAKHA AL ARKHAN
LUNA KAMILA TSABITHA SAKHI
NAWRA RAMADHANIA PUTRI
ASSYABIA RAMIRA PANUNTUN
KAYYISAH NAIRA ALIYA HANUN
SHAQUEENA"""

    results = {}
    for c_name, txt in [("Kelas 4A", text_4a), ("Kelas 4B", text_4b), ("Kelas 4C", text_4c)]:
        students = []
        for line in txt.split("\n"):
            line = line.strip()
            if line:
                students.append({"name": line, "gender": guess_gender(line)})
        results[c_name] = students
    return results

def main():
    final_data = {}
    
    # 1. 2A
    res2a = parse_excel_sheets("data/ABSEN 2A 2026_2027.xlsx")
    # it only has 1 sheet usually, map it to 'Kelas 2A'
    for k, v in res2a.items():
        if v: final_data["Kelas 2A"] = v
        
    # 2. 5D
    res5d = parse_excel_sheets("data/ABSEN SISWA 5D 2026 2027.xlsx")
    for k, v in res5d.items():
        if v: final_data["Kelas 5D"] = v
        
    # 3. Kelas 1 (1A, 1B, 1C)
    res1 = parse_excel_sheets("data/Pembagian_Kelas_1_2026_2027.xlsx")
    for k, v in res1.items():
        if v: final_data[k] = v
        
    # 4. 3C
    res3c = parse_docx("data/Data siswa 3C 26-27.docx")
    final_data.update(res3c)
    
    # 5. 4A, 4B, 4C
    res4 = parse_pdf_text_manual()
    final_data.update(res4)
    
    with open("scripts/final_students.json", "w") as out:
        json.dump(final_data, out, indent=2)

if __name__ == "__main__":
    main()

import openpyxl
import json

f1 = "data/JADWAL 2627 UPDATE.xlsx"
wb = openpyxl.load_workbook(f1, data_only=True)

# Inspect WALAS sheet
sheet_walas = wb["WALAS"]
print("--- SHEET: WALAS ---")
for i, row in enumerate(sheet_walas.iter_rows(values_only=True)):
    row_strs = [str(c).strip() if c is not None else "" for c in row if c is not None]
    if row_strs:
        print(f"Row {i+1}: {row_strs}")

# Inspect PENGGANTI U DESTY
if "PENGGANTI U DESTY" in wb.sheetnames:
    print("\n--- SHEET: PENGGANTI U DESTY ---")
    for i, row in enumerate(wb["PENGGANTI U DESTY"].iter_rows(values_only=True)):
        row_strs = [str(c).strip() if c is not None else "" for c in row if c is not None]
        if row_strs:
            print(f"Row {i+1}: {row_strs}")

# Inspect PENGGANTI U RETNO DAN U ARUM
if "PENGGANTI U RETNO DAN U ARUM" in wb.sheetnames:
    print("\n--- SHEET: PENGGANTI U RETNO DAN U ARUM ---")
    for i, row in enumerate(wb["PENGGANTI U RETNO DAN U ARUM"].iter_rows(values_only=True)):
        row_strs = [str(c).strip() if c is not None else "" for c in row if c is not None]
        if row_strs:
            print(f"Row {i+1}: {row_strs}")

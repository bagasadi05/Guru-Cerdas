import os
import re
import json
import zipfile
import openpyxl
import docx
import pypdf

def guess_gender(name):
    n = name.strip().lower()
    last_word = n.split()[-1] if n else ""
    if last_word in ['putri', 'azzahra', 'ayu', 'nisa', 'zahra', 'dewi', 'sari', 'nur', 'sakhi', 'hanun', 'sakina', 'khalisa', 'pustika', 'azalia', 'utami', 'maheswari', 'rahmani', 'wibowo']:
        return 'Perempuan'
    if last_word in ['putra', 'akbar', 'pratama', 'hidayat', 'ramadhan', 'setiawan', 'prasojo', 'nugroho', 'kamal', 'falah', 'wijaya', 'yusuf', 'wibawa', 'santoso']:
        return 'Laki-laki'
    first_word = n.split()[0] if n else ""
    if first_word in ['aisyah', 'anindita', 'alisha', 'arika', 'azzahra', 'bilqis', 'callista', 'dzakira', 'nafia', 'annisa', 'alika', 'kuntum', 'shevahira', 'tyasayu', 'harumi', 'acintya', 'aifriza', 'ainayya', 'alesha', 'kinarayu', 'gendis', 'adzkiya', 'arsyila', 'elshanum', 'qiandra', 'yurike', 'zizi', 'elmira', 'naila', 'luna', 'nawra', 'assyabia', 'kayyisah', 'shaqueena', 'clarisha', 'chayra', 'adifa', 'adreena', 'aisha', 'almahyra', 'alsyakila', 'aqilla', 'aqila', 'arsya', 'ceisya', 'gladysa', 'khayra', 'myeisha', 'queenzy', 'shakeela', 'zianka', 'adiba', 'bellvania', 'khalifa', 'kayyisah', 'fiona', 'fifi', 'siti', 'dwi', 'tri', 'ratna']:
        return 'Perempuan'
    return 'Laki-laki'

def clean_name(val_str):
    if not val_str: return ""
    val_str = str(val_str).strip()
    val_str = re.sub(r'^\d+[\.\s\-]+', '', val_str).strip()
    return val_str

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = []
    seen = set()
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for row in sheet.iter_rows(values_only=True):
            if not row: continue
            row_strs = [str(cell).strip() if cell is not None else "" for cell in row]
            
            gender = None
            for val in row_strs:
                if val.lower() in ['l', 'laki-laki', 'laki', 'l (laki-laki)']: gender = 'Laki-laki'
                elif val.lower() in ['p', 'perempuan', 'p (perempuan)']: gender = 'Perempuan'
            
            name = ""
            for val in row_strs:
                cleaned = clean_name(val)
                if len(cleaned) > 3 and not re.match(r'^[\d\W]+$', cleaned) and cleaned.lower() not in ['l', 'p', 'laki-laki', 'perempuan', 'l (laki-laki)', 'p (perempuan)', 'nama', 'nama siswa', 'no', 'jenis kelamin']:
                    if len(cleaned) > len(name) and not any(kw in cleaned.upper() for kw in ["KOTA", "DAFTAR", "ABSEN", "KELAS", "SEKOLAH", "MADRASAH", "TAHUN", "WALI", "JUMLAH", "MADIUN", "REKAP", "TEMPAT", "TANGGAL"]):
                        name = cleaned
            
            if name and name not in seen:
                seen.add(name)
                if not gender: gender = guess_gender(name)
                results.append({"name": name, "gender": gender})
    return results

def parse_docx(filepath):
    results = []
    seen = set()
    try:
        doc = docx.Document(filepath)
        for table in doc.tables:
            for row in table.rows:
                row_strs = [cell.text.strip() for cell in row.cells]
                gender = None
                for val in row_strs:
                    if val.lower() in ['l', 'laki-laki', 'laki']: gender = 'Laki-laki'
                    elif val.lower() in ['p', 'perempuan']: gender = 'Perempuan'
                
                name = ""
                for cell in row.cells:
                    cleaned = clean_name(cell.text)
                    if len(cleaned) > 3 and not re.match(r'^[\d\W]+$', cleaned) and cleaned.lower() not in ['l', 'p', 'laki-laki', 'perempuan', 'nama', 'nama siswa', 'no', 'jenis kelamin']:
                        if len(cleaned) > len(name) and not any(kw in cleaned.upper() for kw in ["KOTA", "DAFTAR", "ABSEN", "KELAS", "WALI", "SEKOLAH", "MADRASAH", "TAHUN", "JUMLAH", "MADIUN", "REKAP"]):
                            name = cleaned
                if name and name not in seen:
                    seen.add(name)
                    if not gender: gender = guess_gender(name)
                    results.append({"name": name, "gender": gender})
    except Exception as e:
        print(f"Standard docx parse failed for {filepath}: {e}, using xml fallback...")
        try:
            with zipfile.ZipFile(filepath) as z:
                xml_content = z.read('word/document.xml').decode('utf-8')
                # find text inside <w:t> tags
                text_nodes = re.findall(r'<w:t[^>]*>(.*?)</w:t>', xml_content)
                current_line = []
                for t in text_nodes:
                    cleaned = clean_name(t)
                    if len(cleaned) > 3 and not re.match(r'^[\d\W]+$', cleaned):
                        if not any(kw in cleaned.upper() for kw in ["KOTA", "DAFTAR", "ABSEN", "KELAS", "WALI", "SEKOLAH", "MADRASAH", "TAHUN", "JUMLAH", "MADIUN", "REKAP"]):
                            if cleaned not in seen:
                                seen.add(cleaned)
                                results.append({"name": cleaned, "gender": guess_gender(cleaned)})
        except Exception as e2:
            print(f"XML fallback failed for {filepath}: {e2}")
    return results

def parse_pdf(filepath):
    results = []
    seen = set()
    try:
        reader = pypdf.PdfReader(filepath)
        for page in reader.pages:
            text = page.extract_text()
            lines = text.split('\n')
            for line in lines:
                parts = line.split()
                gender = None
                if len(parts) > 0 and parts[-1].upper() in ['L', 'P']:
                    gender = 'Laki-laki' if parts[-1].upper() == 'L' else 'Perempuan'
                    raw_name = " ".join(parts[:-1])
                else:
                    raw_name = line
                
                cleaned = clean_name(raw_name)
                if len(cleaned) > 3 and not re.match(r'^[\d\W]+$', cleaned):
                    if not any(kw in cleaned.upper() for kw in ["KOTA", "DAFTAR", "ABSEN", "KELAS", "WALI", "SEKOLAH", "MADRASAH", "TAHUN", "JUMLAH", "MADIUN", "REKAP", "TEMPAT", "TANGGAL"]):
                        if cleaned not in seen:
                            seen.add(cleaned)
                            if not gender: gender = guess_gender(cleaned)
                            results.append({"name": cleaned, "gender": gender})
    except Exception as e:
        print(f"Error parsing PDF {filepath}: {e}")
    return results

def main():
    data_dir = "data"
    parsed = {}
    
    file_map = {
        "ABSEN 1A 2026-2027.docx": "Kelas 1A",
        "Absen Kelas 1B.pdf": "Kelas 1B",
        "ABSEN 1C 2627.pdf": "Kelas 1C",
        "ABSEN 2A 2026_2027.xlsx": "Kelas 2A",
        "ABSEN 2B 2627.xlsx": "Kelas 2B",
        "ABSENSI 2C 2026-2027.docx": "Kelas 2C",
        "Data siswa 3C 26-27.docx": "Kelas 3C",
        "Data siswa 4A 2026-2027.docx": "Kelas 4A",
        "ABSENSI KELAS 4B 2627.xlsx": "Kelas 4B",
        "Daftar Siswa Kelas 4C  20262027.docx": "Kelas 4C",
        "Absensi Siswa Kelas 5A 2627.xlsx": "Kelas 5A",
        "ABSENSI KELAS 5B 2526 - Copy.xlsx": "Kelas 5B",
        "DAFTAR PESERTA DIDIK KELAS V-C_T.A 2026-2027.xlsx": "Kelas 5C",
        "ABSEN SISWA 5D 2026 2027.xlsx": "Kelas 5D",
        "Data kelas 6A.xlsx": "Kelas 6A",
        "DAFTAR NAMA KELAS 6B 2026- 2027.xlsx": "Kelas 6B",
        "ABSENSI 6C 2026-2027.xlsx": "Kelas 6C",
        "ABSEN 6D BARU 2026.xlsx": "Kelas 6D"
    }

    for fname, target_class in file_map.items():
        fpath = os.path.join(data_dir, fname)
        if not os.path.exists(fpath):
            print(f"File not found: {fpath}")
            continue
        
        ext = fname.split('.')[-1].lower()
        students = []
        if ext == 'xlsx':
            students = parse_excel(fpath)
        elif ext == 'docx':
            students = parse_docx(fpath)
        elif ext == 'pdf':
            students = parse_pdf(fpath)
        
        print(f"[{target_class}] parsed {len(students)} students from {fname}")
        if students:
            parsed[target_class] = students

    with open("scripts/parsed_all_data.json", "w") as out:
        json.dump(parsed, out, indent=2)

if __name__ == "__main__":
    main()

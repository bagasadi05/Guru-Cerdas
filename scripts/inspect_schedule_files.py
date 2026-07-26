import openpyxl
import os

f1 = "data/JADWAL 2627 UPDATE.xlsx"
f2 = "data/JADWAL KELAS 1-6 (2).xlsx"

for filepath in [f1, f2]:
    print(f"\n==========================================")
    print(f"FILE: {filepath}")
    print(f"==========================================")
    if not os.path.exists(filepath):
        print("File not found!")
        continue
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print("Sheets:", wb.sheetnames)
    for sname in wb.sheetnames:
        sheet = wb[sname]
        print(f"\n--- SHEET: {sname} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ---")
        # Print first 15 non-empty rows
        count = 0
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                row_str = [str(c).strip() if c is not None else "" for c in row[:12]]
                print(row_str)
                count += 1
                if count >= 15:
                    print("... (truncated sheet view)")
                    break
